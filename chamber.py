# from bs4 import BeautifulSun k])

from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module="openpyxl")
import csv

wb = load_workbook(filename='chamber_3.xlsx')

ws = wb['Sheet1 (1)']
# with open('biz.txt', "r") as f:
#     for i in f.readlines():
#         print(i.strip())
#     f.close()
# This code created a useable list with only a few things to fix manually
########################################################
data = []
for i in range(1, ws.max_row):
    if ws[i][0].font.b and ws[i][0].font.sz == 10:
        temp_data = []
        for k in range(i, i+7):
            for d in ws[k]:
                if d.value == None:
                    temp_data.append("None")
                else:
                    temp_data.append(d.value)
        # print("\n")
        data.append(temp_data)


with open("big_chamber_list.csv", "a") as f:
    writer = csv.writer(f, delimiter=",")
    for i in data:
        writer.writerow(i)

########################################################
# this grabs stuff with a placehoder of None for things that are missing
###########################################
# for i in range(1, 400):
#     if ws[i][0].font.b and ws[i][0].font.sz == 10:
#         for k in range(i, i+7):
#             if ws.row_dimensions[k].height > 1:
#                 if ws[k][0].value:
#                     pass
#                     # print(ws[k][0].value)
#         print("\n")
    # print(ws.row_dimensions[i].height)
        # names.append([
        #     ws[i][0].value, 
        #     ws[i+1][0].value, 
        #     ws[i+2][0].value, 
        #     # ws[i+3][0].value, 
        #     # ws[i+4][0].value,
        #     # ws[i+5][0].value,
        #     # ws[i+6][0].value,
        #     # ws[i+6][0].value,
        #     # ws[i+2][5].value,
        #     # ws[i+3][6].value,
        #     ]
        #     )

# for i in names:
#     print(i)


# biz = open('biz.txt', "a")
# for i in names:
#     biz.write(i)
    
    
    
# biz.close()

################################################
        
# print(names)

# Ok this following code will grab most stuff. But some businessses have too liitle info. So it grabs stuff from the next
# business
################################################
# GROUP = "Wholesalers/distributors"
# START, END = 4654, 4699
# data = []

# wb = load_workbook(filename='chamber.xlsx')
# ws = wb['Sheet1']

# for i in range(START, END):
#     if ws[i][0].font.b and ws[i][0].font.sz < 14:
#         temp_data = [] 
#         for k in range(i, i+7):
#             for c in ws[k]:
#                 if c.value:
#                     temp_data.append(c.value)
#         data.append(temp_data)
# with open("big_chamber_list.csv", "a") as f:
#     writer = csv.writer(f, delimiter=",")
#     for i in data:
#         writer.writerow(i + [GROUP])
######################################################

# csv.close()
# t = soup.find_all(lambda tag:tag.name == "td" and 
#         bool(tag.attrs) 
#         and tag.attrs['class'] != ['s0']
#         )

#this can sort of get the blocks of data
# for i in range(0, len(t)):
#     if t[i].attrs['class'] == ['s2']:
#         temp_list = t[i:i+9]
#         for k in range(0, len(temp_list)):
#             if temp_list[k].text.strip() == "Phone:":
#                 print(f"{temp_list[k].text.strip()}{temp_list[k+1].text}")
#         print("\n")

# for i in t:
#     print(i)
# for i in t:
#     if i.text == "McDonald's":
#         print(i)

# for i in t:
#     if 'class' in i.attrs:
#         if i.attrs['class'] == ["s2"]:

