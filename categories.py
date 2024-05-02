from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module="openpyxl")
import csv

wb = load_workbook(filename='chamber_3.xlsx')
ws = wb['Sheet1 (1)']

thing = True
cur_header = None
for i in range(1, ws.max_row):
    if ws[i][0].font:
        if ws[i][0].font.b and ws[i][0].font.sz == 14:
            cur_header = ws[i][0].offset(row=1)
        
            while thing:
                if cur_header.font:
                    if cur_header.font.sz == 14:
                        print(cur_header.value)
                        thing =False
                    else:
                        print(cur_header.value)
                        cur_header = cur_header.offset(row=1)



